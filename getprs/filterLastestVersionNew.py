import re
import openpyxl

DATA_FILE = './Oracle补丁项目反措清单.xlsx'#当前目录下待处理文件
#DATA_FILE = './test.csv'
SHEETNAME='反措清单'
STORED_FILE = './Oracle补丁项目反措清单out.xlsx'#处理结果保存到当前目录下目标文件

#获取版本列表中日期最新的版本
def getnewversion(versionList):
    datelist=[]
    result=""
    for versionline in versionList:
        datelist+=re.findall(r"(?<!SZ)(\d{8}|\d{7}|\d{6})", versionline)#零宽负向断言，排除SZ开头的合同号
    if len(datelist) == 0:#如果未匹配出日期，直接返回
        return "无匹配日期"
    else:
        for index in range(len(datelist)):
            if len(str(datelist[index]))==8:
                result+=""
            elif len(datelist[index])==6:#如果是六位日期，先补齐为8位
                datelist[index]='20'+datelist[index]
            else:#如果是其他位数（7位）的日期，提示
                result+="请检查%s日期格式."%datelist[index]
        return result+datelist[datelist.index(max(datelist))]#返回提示信息+最新版本

def getLatestDate(versionText):
    listVersion = []
    matchV = re.findall(r"((PRS-700U|PRS-7000|PRS7000|PRS700U).+?(?<!SZ)(\d{8}|\d{7}|\d{6}))", versionText,
                        re.MULTILINE)  # 非贪婪（懒惰）匹配，匹配尽量少的内容，以免将两条版本信息匹配为一条
    for matchVersion in matchV:  # findall会匹配所有()分组，取第一个，即最外层()的匹配
        listVersion.append(matchVersion[0])
    return getnewversion(listVersion)

#wbname==即文件名称，sheetname==工作表名称，可以为空，若为空默认第一个工作表
def readwb(wbname,sheetname):
    wb=openpyxl.load_workbook(filename=wbname,read_only=True)
    if (sheetname==""):
        ws=wb.get_active_sheet()
    else:
        ws=wb.get_sheet_by_name(sheetname)
    data=[]
    lastIndex=""
    lastVersion=""
    rowNum=ws.max_row
    colNum=ws.max_column
    for i in range(1,rowNum):
        list=[]
        thisIndex=ws.cell(row=i,column=1).value
        thisVersion=getLatestDate(ws.cell(row=i,column=5).value)
        if lastIndex=="":
            lastIndex = thisIndex
            lastVersion = thisVersion
        elif thisIndex==lastIndex:
            if thisVersion=="无匹配日期":
                print(str(i)+"无匹配日期")
            if thisVersion>lastVersion:
                for j in range(1,colNum):
                    aa = str(ws.cell(row=i,column=j).value)
                    if (aa == "None"):
                        aa = ""
                    list.append(aa)
                data.pop()
                data.append(list)
                lastIndex = thisIndex
                lastVersion = thisVersion
        else:
            for j in range(1, colNum):
                aa = str(ws.cell(row=i,column=j).value)
                if (aa == "None"):
                    aa = ""
                list.append(aa)
            data.append(list)
            lastIndex = thisIndex
            lastVersion = thisVersion

    print (wbname +"-"+sheetname+"- 已成功读取")
    return data

#新建excel
def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

# 写入excel文件中 date 数据，date是list数据类型， fields 表头
def savetoexcel(data,fields,sheetname,wbname):
    print("写入excel：")
    wb=openpyxl.load_workbook(filename=wbname)

    sheet=wb.active
    sheet.title=sheetname

    field=1
    for field in range(1,len(fields)+1):   # 写入表头
        _=sheet.cell(row=1,column=field,value=str(fields[field-1]))

    row1=1
    col1=0
    for row1 in range(2,len(data)+2):  # 写入数据
        for col1 in range(1,len(data[row1-2])+1):
            _=sheet.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    wb.save(filename=wbname)
    print("保存成功")

datas=readwb(DATA_FILE,SHEETNAME)
headerlist=['合同号', '合同名称', '省份', '装置型号', '软件版本', '软件路径', '汇总', '来源']
creatwb(STORED_FILE)
savetoexcel(datas,headerlist,SHEETNAME,STORED_FILE)